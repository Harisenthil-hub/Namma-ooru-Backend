from openpyxl import Workbook
from django.http import HttpResponse
from rest_framework.views import APIView
from .utils import get_filtered_orders
from rest_framework.permissions import IsAdminUser
from openpyxl.styles import Font
from django.db.models import Sum, OuterRef, Count, Q, Max, Subquery
from django.utils import timezone
from django.utils.timezone import localtime
from .models import Address, Customer
from datetime import timedelta
from decimal import Decimal


class AdminOrderExportAPIView(APIView):
    permission_classes = [IsAdminUser]
    
    def get(self,request):
        
        orders = get_filtered_orders(request).prefetch_related('items')
        include_summary = request.query_params.get('summary') == 'true'
        
        wb = Workbook()
        ws_orders = wb.active
        ws_orders.title = 'Orders'
        
        orders_headers = [
            'Order Number',
            'Customer Number',
            'Phone',
            'Alternate Phone',
            'Status',
            'Total Amount',
            'Street',
            'City',
            'Landmark',
            'Pincode',
            'Created At',
        ]
        
        # For Orders sheet
        ws_orders.append(orders_headers)
        
        for cell in ws_orders[1]:
            cell.font = Font(bold=True)
        
        # For Order Items sheet
        ws_items = wb.create_sheet(title='Order Items')
        
        items_headers = [
            'Order Number',
            'Product',
            'Variant Weight',
            'Quantity',
            'Unit Price',
            'Total Price',
        ]
        
        ws_items.append(items_headers)
        
        for cell in ws_items[1]:
            cell.font = Font(bold=True)
        
        for order in orders.iterator(chunk_size=200):
            
            # Orders Sheet
            ws_orders.append([
                order.order_number,
                order.customer.name,
                order.customer.phone_no,
                order.customer.alternate_phone_no,
                order.order_status,
                order.total_amount,
                order.shipping_address.street if order.shipping_address else "",
                order.shipping_address.city if order.shipping_address else "",
                order.shipping_address.landmark if order.shipping_address else "",
                order.shipping_address.pincode if order.shipping_address else "",
                localtime(order.created_at).strftime('%Y-%m-%d %H:%M'),
            ])
            
            # Items Sheet
            
            for item in order.items.all():
                ws_items.append([
                    order.order_number,
                    item.product_name,
                    item.variant_weight,
                    item.quantity,
                    item.unit_price,
                    item.total_price,
                ])
            
        # optional Summary sheet   
        if include_summary:
            
            ws_summary = wb.create_sheet(title='Summary')
            
            
            # Metrics section
            
            total_orders = orders.count()
            confirmed_orders = orders.filter(order_status='confirmed').count()
            pending_orders = orders.filter(order_status='pending').count()
            cancelled_orders = orders.filter(order_status='cancelled').count()
            revenue = orders.filter(order_status='confirmed').aggregate(
                total=Sum('total_amount')
            )['total'] or 0
            pending_value = orders.filter(order_status='pending').aggregate(
                total=Sum('total_amount')
            )['total'] or 0
            cancelled_value = orders.filter(order_status='cancelled_value').aggregate(
                total=Sum('total_amount')
            )['total'] or 0
            conversion_rate = (confirmed_orders / total_orders) * 100 if total_orders else 0
            
            
            
            ws_summary.append(['Metrics', 'Value'])
            ws_summary.append(['Total Orders', total_orders])
            ws_summary.append(['Confirmed Orders', confirmed_orders])
            ws_summary.append(['Pending Orders', pending_orders])
            ws_summary.append(['Cancelled Orders', cancelled_orders])
            ws_summary.append(['Revenue', revenue])
            ws_summary.append(['Pending Value', pending_value])
            ws_summary.append(['Cancelled Value', cancelled_value])
            ws_summary.append(['Conversion Rate', conversion_rate])
            
            
            ws_summary.append([])
            ws_summary.append([])
            
            status = request.query_params.get('status')
            search = request.query_params.get('search')
            date = request.query_params.get('date')
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            quick = request.query_params.get('quick')
            min_amount = request.query_params.get('min_amount')
            max_amount = request.query_params.get('max_amount')
            
            # Filters Section
            
            ws_summary.append(["Filters Used",""])
            ws_summary.append(["Status", status or "All"])
            ws_summary.append(["Search", search or "None"])
            ws_summary.append(["Date", date or "None"])
            ws_summary.append(["Start Date", start_date or "None"])
            ws_summary.append(["End Date", end_date or "None"])
            ws_summary.append(["Quick Filter", quick or "None"])
            ws_summary.append(["Min Amount", min_amount or "None"])
            ws_summary.append(["Max Amount", max_amount or "None"])
            
             
            ws_summary.append([])
        
            ws_summary.append(["Export Generated At", localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")])
            
            
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response['Content-Disposition'] = 'attachment; filename="orders.xlsx"'
        
        wb.save(response)
        
        return response
    


class AdminCustomerExportAPIView(APIView):
    permission_classes = [IsAdminUser]
    
    
    def get_queryset(self,request):
        
        
        search = request.GET.get('search','').strip()
        customer_type = request.GET.get('customer_type')
        min_orders = request.GET.get('min_orders')
        min_spent = request.GET.get('min_spent')
        date = request.GET.get('date')
        start_date = request.GET.get('start_date')
        quick = request.query_params.get('quick')
        end_date = request.GET.get('end_date')
        
        latest_address = Address.objects.filter(
            customer=OuterRef('pk')
        ).order_by('-created_at')
        
        qs = (
            Customer.objects
            .annotate(
                total_orders=Count('orders',distinct=True),
                
                pending_orders=Count(
                    'orders',
                    filter=Q(orders__order_status='pending')
                ),
                
                confirmed_orders=Count(
                    'orders',
                    filter=Q(orders__order_status='confirmed')
                ),
                
                cancelled_orders=Count(
                    'orders',
                    filter=Q(orders__order_status='cancelled')
                ),
                
                confirmed_spent=Sum(
                    'orders__total_amount',
                    filter=Q(orders__order_status='confirmed')
                ),
                
                last_order_at=Max('orders__created_at'),
                
                street=Subquery(
                    latest_address.values('street')[:1]
                ),
                
                city=Subquery(
                    latest_address.values('city')[:1]
                ),
                
                landmark=Subquery(
                    latest_address.values('landmark')[:1]
                ),
                
                pincode=Subquery(
                    latest_address.values('pincode')[:1]
                ),   
            ).order_by('-last_order_at')
        )
        
        
        if search:
           qs = qs.filter(
               Q(phone_no__icontains=search) |
               Q(name__icontains=search)
           )
           
        
        if date:
            qs = qs.filter(
                last_ordet_at__date=date
            )
            
        if start_date and end_date:
            qs = qs.filter(last_order_at__range=[start_date,end_date])
            
        
        if quick:
            now = timezone.now()
            
            if quick == 'today':
                qs = qs.filter(last_order_at__date=now.date())
                
            elif quick == 'yesterday':
                qs = qs.filter(last_order_at__date=(now - timedelta(day=1)).date())
                
            elif quick == 'last7days':
                qs = qs.filter(last_order_at__gte=now - timedelta(days=7))
                
            elif quick == 'last30days':
                qs = qs.filter(last_order_at__gte=now - timedelta(days=30))
                
                
        if min_orders:
            qs = qs.filter(
                total_orders__gte=min_orders
            )
            
        if min_spent:
            qs = qs.filter(
                confirmed_spent__gte=min_spent
            )
            
            
        return qs
        
    def get(self,request):
        
        customers = self.get_queryset(request)
        include_summary = request.GET.get('summary')
    
        wb = Workbook()
        ws = wb.active
        ws.title = 'Customers'
        
        
        headers = [
            'Customer Name',
            'Phone',
            'Alternate Phone'
            'Street',
            'City',
            'Landmark',
            'Pincode',
            'Total Orders',
            'Confirmed Orders',
            'Pending Orders',
            'Cancelled Orders',
            'Confirmed Revenue'
            'Avg Order Value',
            'Confirmation Rate %',
            'Cancellation Rate %',
            'Last Order Date',
        ]
        
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            
            
        total_customers = 0
        total_orders_sum = 0
        pending_orders_sum = 0
        confirmed_orders_sum = 0
        cancelled_orders_sum = 0
        total_revenue = Decimal('0.00')
        
        for customer in customers.iterator(chunk_size=200):
            
            total_orders = customer.total_orders or 0
            pending = customer.pending_orders or 0
            confirmed = customer.confirmed_orders or 0
            cancelled = customer.cancelled_orders or 0
            revenue = customer.confirmed_spent or Decimal('0.00')
            
            avg_value = (
                revenue / confirmed
                if confirmed > 0 else Decimal('0.00')
            )
            
            cancel_rate = (
                (cancelled / total_orders) * 100
                if total_orders > 0 else 0
            )
            
            confirm_rate = (
                (confirmed / total_orders) * 100
                if total_orders > 0 else 0
            )
            
            
            ws.append([
                customer.name,
                customer.phone_no,
                customer.alternate_phone_no,
                customer.street,
                customer.city,
                customer.landmark,
                customer.pincode,
                total_orders,
                pending,
                confirmed,
                cancelled,
                Decimal(str(revenue)),
                Decimal(str(avg_value)),
                round(confirm_rate,2),
                round(cancel_rate,2),
                localtime(customer.last_order_at).strftime('%Y-%m-%d %H:%M')
                if customer.last_order_at else ""
            ])
            
            total_customers += 1
            total_orders_sum += total_orders
            pending_orders_sum += pending
            confirmed_orders_sum += confirmed
            cancelled_orders_sum += cancelled
            total_revenue += revenue
            
            
        # Summary sheet
        summary = wb.create_sheet('Summary')
        
        summary.append(['Metric','Value'])
        
        summary.append(['Total Customers',total_customers])
        summary.append(['Total Orders',total_orders_sum])
        summary.append(['Pending Orders',pending_orders_sum])
        summary.append(['Confirmed Orders',confirmed_orders_sum])
        summary.append(['Cancelled Orders',cancelled_orders_sum])
        summary.append(['Total Confirmed Revenue',Decimal(str(total_revenue or 0))])
        
        avg_customer_spend = (
            total_revenue / total_customers
            if total_customers > 0 else 0
        )
        
        summary.append(['Average Customer Spend',Decimal(str(avg_customer_spend or 0))])
        
        overall_cancel_rate = (
            (cancelled_orders_sum / total_orders_sum) * 100
            if total_orders_sum > 0 else 0
        )
        
        overall_confirm_rate = (
            (confirmed_orders_sum / total_orders_sum) * 100
            if total_orders_sum > 0 else 0
        )
        
        summary.append(['Overall Cancellation Rate %',round(overall_cancel_rate,2)])
        summary.append(['Overall Confirmation Rate %',round(overall_confirm_rate,2)])
        
        
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f'customers_export_{timezone.now().date()}.xlsx'
        
        response['Content-Disposition'] = (
            f'attachment; filename="{filename}"'
        )
        
        
        wb.save(response)
        
        return response
        
        
            
            
            