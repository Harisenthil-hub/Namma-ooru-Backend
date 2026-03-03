from openpyxl import Workbook
from openpyxl.styles import Font
from django.utils.timezone import localtime


def generate_product_export(products, request, variant_mode, include_summary):
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'
    
    headers = [
        'Product ID',
        'Product Name',
        'Category',
        'Product Active',
        'Product Update at',
        'Variant Weight',
        'Variant Price',
        'Variant Offer Price',
        'Variant Active',
        'Variant Updated at',
    ]
    
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    for product in products:
        if variant_mode == 'updated_only':
            variants = product.variants.filter(
                update_at__date=product.updated_at.date()
            )
        else:
            variants = product.variants.all()
    
        if variants.exists():
            for variant in variants:
                ws.append([
                    product.id,
                    product.name,
                    product.category.name if product.category else "",
                    product.is_active,
                    localtime(product.updated_at).strftime('%Y-%m-%d %H:%M'),
                    variant.weight,
                    variant.price,
                    variant.offer_price or '',
                    variant.is_active,
                    localtime(variant.updated_at).strftime('%y-%m-%d %H:%M'),
                ])
        else:
            ws.append([
                    product.id,
                    product.name,
                    product.category.name if product.category else "",
                    product.is_active,
                    localtime(product.updated_at).strftime('%Y-%m-%d %H:%M'),
                    "" , "", "", "", ""
            ])
            
    # Optional Summary Sheet
    if include_summary:
        summary_sheet = wb.create_sheet('Summary')
        summary_sheet.append(['Total Products', products.count()])
        summary_sheet.append(['Total Variants', sum(p.variants.count() for p in products)])
        summary_sheet.append(['Filters',str(request.GET.dict())])
        
    return wb
        